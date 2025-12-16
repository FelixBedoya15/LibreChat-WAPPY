import openpyxl
import os

# File configuration
filename = '2026  agenda.xlsm'
sheet_name = 'Enero 2026'
target_value = '1466'
replacement_value = '20000'

def modify_excel():
    if not os.path.exists(filename):
        print(f"Error: File '{filename}' not found.")
        return

    print(f"Loading {filename}...")
    try:
        # Load the workbook, keeping vba macros
        wb = openpyxl.load_workbook(filename, keep_vba=True)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return

    if sheet_name not in wb.sheetnames:
        print(f"Error: Sheet '{sheet_name}' not found in the workbook.")
        print(f"Available sheets: {wb.sheetnames}")
        return

    ws = wb[sheet_name]
    print(f"Processing sheet: {sheet_name}")

    modifications = 0
    
    # Iterate through all cells in the worksheet
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == 'f':  # Check if cell contains a formula
                formula = str(cell.value)
                if target_value in formula:
                    # Replace the target value in the formula
                    new_formula = formula.replace(target_value, replacement_value)
                    cell.value = new_formula
                    modifications += 1
                    # print(f"Modified cell {cell.coordinate}: {formula} -> {new_formula}")

    print(f"Total modifications made: {modifications}")

    if modifications > 0:
        try:
            print(f"Saving {filename}...")
            wb.save(filename)
            print("File saved successfully.")
        except Exception as e:
            print(f"Error saving file: {e}")
    else:
        print("No changes made.")

if __name__ == "__main__":
    modify_excel()
