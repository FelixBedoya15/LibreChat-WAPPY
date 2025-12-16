import openpyxl
from copy import copy
import calendar
from datetime import datetime
import locale

# Set locale to Spanish for month/day names
try:
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
except locale.Error:
    print("Warning: Spanish locale not found. Using default.")

filename = '2026  agenda.xlsm'
source_sheet_name = 'Enero 2026'

def create_months():
    print(f"Loading {filename}...")
    try:
        wb = openpyxl.load_workbook(filename, keep_vba=True)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return

    if source_sheet_name not in wb.sheetnames:
        print(f"Error: Source sheet '{source_sheet_name}' not found.")
        return

    source_ws = wb[source_sheet_name]
    
    # Spanish month names (1-indexed for convenience, 0 is empty)
    month_names = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", 
                   "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

    # Loop from February (2) to December (12)
    for month in range(2, 13):
        new_sheet_name = f"{month_names[month]} 2026"
        
        # Check if sheet already exists
        if new_sheet_name in wb.sheetnames:
            print(f"Sheet '{new_sheet_name}' already exists. Skipping creation (updating dates only).")
            target_ws = wb[new_sheet_name]
        else:
            print(f"Creating sheet: {new_sheet_name}")
            target_ws = wb.copy_worksheet(source_ws)
            target_ws.title = new_sheet_name

        # Update dates for days 1 to 31
        # Pattern: Day 1 is in Col B (2), Day 2 in Col E (5), ... Col = 2 + (day-1)*3
        
        num_days_in_month = calendar.monthrange(2026, month)[1]
        
        for day in range(1, 32):
            col_idx = 2 + (day - 1) * 3
            cell = target_ws.cell(row=1, column=col_idx)
            
            if day <= num_days_in_month:
                date_obj = datetime(2026, month, day)
                cell.value = date_obj
                # Apply Spanish date format: "sábado, 1 de febrero de 2026"
                # Excel format code roughly: [$-es-ES]dddd, d "de" mmmm "de" yyyy
                cell.number_format = '[$-es-ES]dddd, d "de" mmmm "de" yyyy'
            else:
                # Clear invalid days (e.g., Feb 30)
                cell.value = None

    print(f"Saving {filename}...")
    try:
        wb.save(filename)
        print("File saved successfully.")
    except Exception as e:
        print(f"Error saving file: {e}")

if __name__ == "__main__":
    create_months()
