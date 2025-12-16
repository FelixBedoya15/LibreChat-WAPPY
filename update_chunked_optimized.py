import openpyxl
from copy import copy
import calendar
from datetime import datetime
import locale
import sys
import gc

# Set locale
try:
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
except locale.Error:
    print("Warning: Spanish locale not found. Using default.")

filename = '2026  agenda.xlsm'
source_sheet_name = 'Enero 2026'

def process_single_sheet(target_sheet_name, create_if_missing=False):
    print(f"--- Processing {target_sheet_name} ---")
    try:
        # Load workbook
        wb = openpyxl.load_workbook(filename, keep_vba=True)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return False

    sheet_modified = False

    # 1. Create/Get Sheet
    if target_sheet_name not in wb.sheetnames:
        if create_if_missing:
            if source_sheet_name in wb.sheetnames:
                print(f"Creating {target_sheet_name}...")
                source_ws = wb[source_sheet_name]
                ws = wb.copy_worksheet(source_ws)
                ws.title = target_sheet_name
                sheet_modified = True
            else:
                print(f"Error: Source sheet {source_sheet_name} not found.")
                return False
        else:
            print(f"Sheet {target_sheet_name} not found and create_if_missing=False.")
            return False
    else:
        ws = wb[target_sheet_name]

    # 2. Update Dates (if it's a monthly sheet)
    if "2026" in target_sheet_name and target_sheet_name != "Enero 2026":
        try:
            month_str = target_sheet_name.split(" ")[0]
            month_names = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", 
                           "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
            if month_str in month_names:
                month = month_names.index(month_str)
                print(f"Updating dates for {month_str}...")
                num_days_in_month = calendar.monthrange(2026, month)[1]
                for day in range(1, 32):
                    col_idx = 2 + (day - 1) * 3
                    cell = ws.cell(row=1, column=col_idx)
                    
                    expected_value = None
                    if day <= num_days_in_month:
                        expected_value = datetime(2026, month, day)
                    
                    # Only update if different to avoid unnecessary saves if re-running
                    # (Simple check, might not be perfect for objects, but good enough)
                    if cell.value != expected_value:
                        cell.value = expected_value
                        if expected_value:
                            cell.number_format = '[$-es-ES]dddd, d "de" mmmm "de" yyyy'
                        sheet_modified = True
        except ValueError:
            pass

    # 3. Update Formulas
    print(f"Scanning formulas in {target_sheet_name}...")
    formulas_updated = 0
    # Use iter_rows for efficiency
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == 'f':
                formula = str(cell.value)
                new_formula = formula
                
                if '5000' in formula:
                    new_formula = new_formula.replace('5000', '20000')
                if '1466' in formula:
                    new_formula = new_formula.replace('1466', '20000')
                
                if new_formula != formula:
                    cell.value = new_formula
                    formulas_updated += 1
                    sheet_modified = True
    
    print(f"Formulas updated: {formulas_updated}")

    # 4. Save if modified
    if sheet_modified:
        print(f"Saving changes to {target_sheet_name}...")
        try:
            wb.save(filename)
            print("Saved.")
        except Exception as e:
            print(f"Error saving: {e}")
            return False
    else:
        print("No changes needed.")
    
    # Cleanup
    del wb
    gc.collect()
    return True

def main():
    month_names = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", 
                   "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

    # Process Enero first
    process_single_sheet("Enero 2026", create_if_missing=False)

    # Process Feb-Dec
    for month in range(2, 13):
        sheet_name = f"{month_names[month]} 2026"
        process_single_sheet(sheet_name, create_if_missing=True)

if __name__ == "__main__":
    main()
