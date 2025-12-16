import openpyxl
from copy import copy
import calendar
from datetime import datetime
import locale

# Set locale to Spanish for month/day names
try:
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
except locale.Error:
    print("Warning: Spanish locale not found. Using default.")

filename = '2026  agenda.xlsm'
source_sheet_name = 'Enero 2026'

def update_all():
    print(f"Loading {filename}...")
    try:
        # read_only=False is default. keep_vba=True is essential.
        wb = openpyxl.load_workbook(filename, keep_vba=True)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return

    if source_sheet_name not in wb.sheetnames:
        print(f"Error: Source sheet '{source_sheet_name}' not found.")
        return

    source_ws = wb[source_sheet_name]
    
    month_names = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", 
                   "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

    # 1. Ensure all monthly sheets exist and have correct dates
    print("Checking/Creating monthly sheets...")
    for month in range(1, 13):
        sheet_name = f"{month_names[month]} 2026"
        
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            print(f"Creating sheet: {sheet_name}")
            ws = wb.copy_worksheet(source_ws)
            ws.title = sheet_name

        # Update dates for days 1 to 31
        num_days_in_month = calendar.monthrange(2026, month)[1]
        
        for day in range(1, 32):
            col_idx = 2 + (day - 1) * 3
            cell = ws.cell(row=1, column=col_idx)
            
            if day <= num_days_in_month:
                date_obj = datetime(2026, month, day)
                cell.value = date_obj
                cell.number_format = '[$-es-ES]dddd, d "de" mmmm "de" yyyy'
            else:
                cell.value = None

    # 2. Update formulas in ALL sheets
    print("Updating formulas in all sheets...")
    total_modifications = 0
    
    for ws in wb.worksheets:
        print(f"Scanning sheet: {ws.title}")
        sheet_mods = 0
        # Optimization: Use ws.iter_rows() with values_only=False (default)
        # We only care about cells with formulas.
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f':
                    formula = str(cell.value)
                    modified = False
                    
                    if '5000' in formula:
                        formula = formula.replace('5000', '20000')
                        modified = True
                    
                    if '1466' in formula:
                        formula = formula.replace('1466', '20000')
                        modified = True
                        
                    if modified:
                        cell.value = formula
                        sheet_mods += 1
                        total_modifications += 1
        print(f"  - Modifications in {ws.title}: {sheet_mods}")

    print(f"Total formula modifications: {total_modifications}")

    print(f"Saving {filename}...")
    try:
        wb.save(filename)
        print("File saved successfully.")
    except Exception as e:
        print(f"Error saving file: {e}")

if __name__ == "__main__":
    update_all()
