import openpyxl

filename = '2026  agenda.xlsm'
sheet_name = 'Enero 2026'

try:
    wb = openpyxl.load_workbook(filename, data_only=True) # data_only=True to see values, not formulas
    if sheet_name not in wb.sheetnames:
        print(f"Sheet {sheet_name} not found.")
        exit()
    
    ws = wb[sheet_name]
    
    print("Inspecting first 10 rows and 10 columns:")
    for row in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=10):
        row_data = []
        for cell in row:
            row_data.append(f"{cell.coordinate}: {cell.value}")
        print(" | ".join(row_data))

except Exception as e:
    print(f"Error: {e}")
