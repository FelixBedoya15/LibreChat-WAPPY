import openpyxl
from copy import copy
import calendar
from datetime import datetime
import locale
import sys

# Set locale to Spanish for month/day names
try:
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
except locale.Error:
    print("Warning: Spanish locale not found. Using default.")

filename = '2026  agenda.xlsm'
source_sheet_name = 'Enero 2026'

def process_sheet(sheet_name, create_from=None):
    print(f"Processing {sheet_name}...")
    try:
        wb = openpyxl.load_workbook(filename, keep_vba=True)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return

    # Create if needed
    if sheet_name not in wb.sheetnames:
        if create_from and create_from in wb.sheetnames:
            print(f"  Creating {sheet_name} from {create_from}")
            source_ws = wb[create_from]
            ws = wb.copy_worksheet(source_ws)
            ws.title = sheet_name
        else:
            print(f"  Error: Cannot create {sheet_name}, source {create_from} not found.")
            return
    else:
        print(f"  {sheet_name} exists.")
        ws = wb[sheet_name]

    # Update Dates (only for monthly sheets)
    if "2026" in sheet_name and sheet_name != "Enero 2026": # Skip Enero date update as it's the source
        try:
            month_str = sheet_name.split(" ")[0]
            month_names = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", 
                           "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
            month = month_names.index(month_str)
            
            print(f"  Updating dates for {month_str}...")
            num_days_in_month = calendar.monthrange(2026, month)[1]
            for day in range(1, 32):
                col_idx = 2 + (day - 1) * 3
                cell = ws.cell(row=1, column=col_idx)
                if day <= num_days_in_month:
                    date_obj = datetime(2026, month, day)
                    cell.value = date_obj
                    cell.number_format = '[$-es-ES]dddd, d "de" mmmm "de" yyyy'
                else:
                    cell.value = None
        except ValueError:
            pass # Not a standard month sheet name

    # Update Formulas
    print(f"  Updating formulas in {sheet_name}...")
    modifications = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == 'f':
                formula = str(cell.value)
                modified = False
                if '5000' in formula:
                    formula = formula.replace('5000', '20000')
                    modified = True
                if '1466' in formula:
                    formula = formula.replace('1466', '20000')
                    modified = True
                if modified:
                    cell.value = formula
                    modifications += 1
    print(f"  Modifications: {modifications}")

    print(f"  Saving {filename}...")
    wb.save(filename)
    print("  Saved.")

def main():
    month_names = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", 
                   "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

    # Process Enero first (just formula update)
    process_sheet("Enero 2026")

    # Process Feb-Dec (Create + Date + Formula)
    for month in range(2, 13):
        sheet_name = f"{month_names[month]} 2026"
        process_sheet(sheet_name, create_from="Enero 2026")

if __name__ == "__main__":
    main()
